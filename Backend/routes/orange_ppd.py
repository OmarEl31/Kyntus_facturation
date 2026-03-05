# Backend/routes/orange_ppd.py
from __future__ import annotations

import csv
import io
import re
import unicodedata
import uuid
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import text
from sqlalchemy.orm import Session

from database.connection import get_db
from models.raw_orange_ppd_import import RawOrangePpdImport
from models.raw_orange_ppd_row import RawOrangePpdRow
from models.raw_orange_ppd_pivot_row import RawOrangePpdPivotRow

router = APIRouter(prefix="/api/orange-ppd", tags=["orange-ppd"])


# --------------------
# Normalisation helpers
# --------------------
def _norm(s: str) -> str:
    s = (s or "").replace("\ufeff", "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def _normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k, v in (raw or {}).items():
        if not k:
            continue
        nk = _norm(k)
        out[nk] = v.strip() if isinstance(v, str) else v
    return out


def _pick(h: dict[str, Any], *keys: str) -> str | None:
    for k in keys:
        v = h.get(k)
        if v is None:
            continue
        vv = str(v).strip()
        if vv:
            return vv
    return None


def _parse_decimal(v: str | None) -> Decimal | None:
    if v is None:
        return None
    raw = str(v).replace(" ", "").replace("\u00a0", "").replace("€", "").replace(",", ".").strip()
    if raw == "":
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _norm_ot(v: str | None) -> str | None:
    if not v:
        return None
    s = re.sub(r"\s+", "", str(v).strip())
    if not s:
        return None
    if s.isdigit():
        s2 = s.lstrip("0")
        return s2 if s2 != "" else "0"
    return s


def _resolve_import_id(db: Session, import_id: str | None) -> str | None:
    if import_id:
        return import_id
    latest = (
        db.query(RawOrangePpdImport)
        .order_by(RawOrangePpdImport.imported_at.desc(), RawOrangePpdImport.import_id.desc())
        .first()
    )
    return latest.import_id if latest else None


# --------------------
# Excel helpers
# --------------------
_HDR_REQUIRED = {"commande", "releve", "montant_brut", "montant_majore"}

_HDR_ALIASES = {
    "commande": {"commande"},
    "gdf": {"gdf", "n_gdf", "no_gdf", "n__gdf"},
    "releve": {"releve"},
    "attachement": {"attach_element", "attache_element", "attach", "attachement"},
    "debut_trvx": {"debut_tvx", "debut_trvx", "debut"},
    "fin_trvx": {"fin_tvx", "fin_trvx", "fin"},
    "montant_brut": {"montant_brut"},
    "montant_majore": {"montant_majore", "montant_majoré", "montant_major"},
}


def _cell_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _find_header_row_and_map(ws) -> tuple[int, dict[str, int]]:
    max_scan = min(ws.max_row or 0, 150)
    max_col = min(ws.max_column or 0, 120)

    for r in range(1, max_scan + 1):
        values = [_cell_str(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        normed = [_norm(v) for v in values]

        found: dict[str, int] = {}
        for key, aliases in _HDR_ALIASES.items():
            for idx, name in enumerate(normed, start=1):
                if name in aliases:
                    found[key] = idx
                    break

        if _HDR_REQUIRED.issubset(found.keys()):
            return r, found

    raise HTTPException(
        status_code=400,
        detail="Header PPD introuvable. Je cherche: Commande, Relevé, Montant brut, Montant majoré.",
    )


def _extract_ppd_meta(ws) -> tuple[str | None, str | None]:
    ppd_num = None
    objet = None

    max_r = min(ws.max_row or 0, 60)
    max_c = min(ws.max_column or 0, 40)

    lines: list[str] = []
    for r in range(1, max_r + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            v = _cell_str(ws.cell(row=r, column=c).value)
            if v:
                row_vals.append(v)
        if row_vals:
            lines.append(" ".join(row_vals))

    blob = "\n".join(lines)

    m = re.search(r"\bPPD\s*n[°o]?\s*[:\-]?\s*([0-9]+)\b", blob, flags=re.IGNORECASE)
    if m:
        ppd_num = m.group(1).strip()

    m2 = re.search(r"\bObjet\s*[:\-]\s*(.+)", blob, flags=re.IGNORECASE)
    if m2:
        objet = m2.group(1).strip()

    return ppd_num, objet


def _stop_line(cmd: str, rel: str, mb: Decimal | None, mm: Decimal | None) -> bool:
    return (cmd.strip() == "") and (rel.strip() == "") and mb is None and mm is None


def _is_xlsx(filename: str | None, content: bytes) -> bool:
    if filename and filename.lower().endswith(".xlsx"):
        return True
    return len(content) >= 2 and content[0:2] == b"PK"


# --------------------
# CSV import (phase 1)
# --------------------
def _import_csv_ppd(db: Session, file: UploadFile, content: bytes, imported_by: str | None):
    txt = content.decode("utf-8-sig", errors="ignore")
    if not txt.strip():
        raise HTTPException(status_code=400, detail="Fichier vide")

    first = txt.splitlines()[0] if txt.splitlines() else ""
    delimiter = ";" if first.count(";") >= first.count(",") else ","

    reader = csv.DictReader(io.StringIO(txt), delimiter=delimiter)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="En-têtes CSV introuvables")

    new_import_id = str(uuid.uuid4())

    imp = RawOrangePpdImport(
        import_id=new_import_id,
        filename=file.filename,
        imported_by=imported_by,
        row_count=0,
    )
    db.add(imp)
    db.flush()

    rows: list[RawOrangePpdRow] = []
    pivots: list[RawOrangePpdPivotRow] = []

    for idx, raw in enumerate(reader, start=1):
        h = _normalize_row(raw)

        etiquette = _pick(h, "etiquettes_de_lignes", "etiquettes_de_ligne", "etiquettes", "etiquette_lignes")
        somme_kyntus = _parse_decimal(_pick(h, "somme_de_kyntus", "somme_kyntus", "total_kyntus"))

        numero_ot_raw = _pick(h, "n_ot", "numero_ot", "num_ot", "ot", "ot_key", "num_ot_orange")
        numero_ot = _norm_ot(numero_ot_raw)

        row_id = f"{new_import_id}:{idx}"

        if not numero_ot and (etiquette or somme_kyntus is not None):
            pivots.append(
                RawOrangePpdPivotRow(
                    row_id=row_id,
                    import_id=new_import_id,
                    etiquette_lignes=etiquette,
                    somme_kyntus=somme_kyntus,
                )
            )
            continue

        if not numero_ot:
            continue

        numero_ppd = _pick(h, "n_ppd", "numero_ppd", "ppd")

        rows.append(
            RawOrangePpdRow(
                row_id=row_id,
                import_id=new_import_id,
                contrat=_pick(h, "contrat"),
                numero_flux_pidi=_pick(h, "n_de_flux_pidi", "numero_flux_pidi", "flux_pidi"),
                type_pidi=_pick(h, "type"),
                statut=_pick(h, "statut"),
                nd=_pick(h, "nd"),
                code_secteur=_pick(h, "code_secteur"),
                numero_ot=numero_ot,
                numero_att=_pick(h, "n_att", "numero_att", "n_att_"),
                oeie=_pick(h, "oeie"),
                code_gestion_chantier=_pick(h, "code_gestion_chantier"),
                agence=_pick(h, "agence"),
                code_postal=_pick(h, "code_postal"),
                code_insee=_pick(h, "code_insee"),
                entreprise=_pick(h, "entreprise"),
                code_gpc=_pick(h, "code_gpc"),
                code_etr=_pick(h, "code_etr"),
                chef_equipe=_pick(h, "chef_d_equipe", "chef_dequipe"),
                ui=_pick(h, "ui"),
                numero_ppd=numero_ppd,
                act_prod=_pick(h, "act_prod"),
                numero_as=_pick(h, "n_as", "numero_as"),
                centre=_pick(h, "centre"),
                date_debut=_pick(h, "date_debut"),
                date_fin=_pick(h, "date_fin"),
                numero_cac=_pick(h, "n_cac", "numero_cac"),
                commentaire_interne=_pick(h, "commentaire_interne"),
                commentaire_oeie=_pick(h, "commentaire_oeie"),
                commentaire_attelem=_pick(h, "commentaire_attelem"),
                motif_facturation_degradee=_pick(h, "motif_facturation_degradee", "motif_de_facturation_degradee"),
                categorie=_pick(h, "categorie"),
                charge_affaire=_pick(h, "charge_affaire", "charge_d_affaire"),
                cause_acqui_rejet=_pick(h, "cause_acqui_rejet"),
                commentaire_acqui_rejet=_pick(h, "comment_acqui_rejet", "commentaire_acqui_rejet"),
                attachement_cree=_pick(h, "attachement_cree"),
                derniere_saisie=_pick(h, "derniere_saisie"),
                attachement_definitif=_pick(h, "attachement_definitif"),
                attachement_valide=_pick(h, "attachement_valide"),
                pointe_ppd=_pick(h, "pointe_ppd"),
                planification_ot=_pick(h, "planification_ot"),
                validation_interventions=_pick(h, "validation_interventions", "validation_des_interventions"),
                bordereau=_pick(h, "bordereau"),
                ht=_parse_decimal(_pick(h, "ht", "total_ht")),
                bordereau_sst=_pick(h, "bordereau_sst"),
                ht_sst=_parse_decimal(_pick(h, "ht_sst")),
                marge=_pick(h, "marge"),
                coeff=_pick(h, "coeff"),
                kyntus=_pick(h, "kyntus"),
                liste_articles=_pick(h, "liste_articles", "liste_des_articles"),
                encours=_pick(h, "encours"),
            )
        )

    if not rows and not pivots:
        raise HTTPException(status_code=400, detail="Aucune ligne exploitable")

    imp.row_count = len(rows)

    if rows:
        db.bulk_save_objects(rows)
    if pivots:
        db.bulk_save_objects(pivots)

    return {
        "ok": True,
        "rows": len(rows),
        "pivots": len(pivots),
        "count": len(rows),
        "message": "Import Orange PPD (CSV) terminé",
        "importId": new_import_id,
        "import_id": new_import_id,
    }


# --------------------
# Excel import (phase 2)
# --------------------
def _import_excel_ppd(db: Session, file: UploadFile, content: bytes, imported_by: str | None, sheet: str | None):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl manquant: {e}")

    wb = load_workbook(io.BytesIO(content), data_only=True)

    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif "PPDATEL" in wb.sheetnames:
        ws = wb["PPDATEL"]
    else:
        ws = wb[wb.sheetnames[0]]

    sheet_name = ws.title
    header_row, col = _find_header_row_and_map(ws)
    ppd_num, objet = _extract_ppd_meta(ws)

    new_import_id = str(uuid.uuid4())

    db.execute(
        text("""
            INSERT INTO canonique.orange_ppd_excel_imports(import_id, filename, sheet_name, imported_by, row_count)
            VALUES (:import_id, :filename, :sheet_name, :imported_by, 0)
        """),
        {
            "import_id": new_import_id,
            "filename": file.filename,
            "sheet_name": sheet_name,
            "imported_by": imported_by,
        },
    )

    data_start = header_row + 1
    max_row = ws.max_row or data_start

    batch: list[dict[str, Any]] = []
    empty_streak = 0
    inserted = 0

    def flush():
        nonlocal inserted
        if not batch:
            return
        db.execute(
            text("""
                INSERT INTO canonique.orange_ppd_excel_rows(
                    row_id, import_id, ppd_num, objet, commande, gdf, releve, attachement,
                    debut_trvx, fin_trvx, montant_brut, montant_majore, sheet_name
                )
                VALUES (
                    :row_id, :import_id, :ppd_num, :objet, :commande, :gdf, :releve, :attachement,
                    :debut_trvx, :fin_trvx, :montant_brut, :montant_majore, :sheet_name
                )
                ON CONFLICT (row_id) DO NOTHING
            """),
            batch,
        )
        inserted += len(batch)
        batch.clear()

    for r in range(data_start, max_row + 1):
        commande = _cell_str(ws.cell(row=r, column=col["commande"]).value)
        releve = _cell_str(ws.cell(row=r, column=col["releve"]).value)

        mb = _parse_decimal(_cell_str(ws.cell(row=r, column=col["montant_brut"]).value))
        mm = _parse_decimal(_cell_str(ws.cell(row=r, column=col["montant_majore"]).value))

        if _stop_line(commande, releve, mb, mm):
            empty_streak += 1
            if empty_streak >= 15:
                break
            continue
        empty_streak = 0

        gdf = _cell_str(ws.cell(row=r, column=col["gdf"]).value) if "gdf" in col else None
        att = _cell_str(ws.cell(row=r, column=col["attachement"]).value) if "attachement" in col else None
        d1 = _cell_str(ws.cell(row=r, column=col["debut_trvx"]).value) if "debut_trvx" in col else None
        d2 = _cell_str(ws.cell(row=r, column=col["fin_trvx"]).value) if "fin_trvx" in col else None

        row_id = f"{new_import_id}:{r}"
        batch.append(
            {
                "row_id": row_id,
                "import_id": new_import_id,
                "ppd_num": ppd_num,
                "objet": objet,
                "commande": commande or None,
                "gdf": gdf or None,
                "releve": releve or None,
                "attachement": att or None,
                "debut_trvx": d1 or None,
                "fin_trvx": d2 or None,
                "montant_brut": mb,
                "montant_majore": mm,
                "sheet_name": sheet_name,
            }
        )

        if len(batch) >= 2000:
            flush()

    flush()

    db.execute(
        text("UPDATE canonique.orange_ppd_excel_imports SET row_count=:n WHERE import_id=:import_id"),
        {"n": int(inserted), "import_id": new_import_id},
    )

    return {
        "ok": True,
        "message": "Import Orange PPD (XLSX) terminé",
        "import_id": new_import_id,
        "importId": new_import_id,
        "rows": int(inserted),
        "count": int(inserted),
        "sheet_name": sheet_name,
        "ppd_num": ppd_num,
        "objet": objet,
        "header_row": header_row,
    }


# --------------------
# Import UNIQUE (CSV ou XLSX)
# --------------------
@router.post("/import")
async def import_orange_ppd(
    file: UploadFile = File(...),
    imported_by: str | None = Query(None),
    sheet: str | None = Query(None),
    db: Session = Depends(get_db),
):
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Fichier vide")

        payload = (
            _import_excel_ppd(db, file, content, imported_by, sheet)
            if _is_xlsx(file.filename, content)
            else _import_csv_ppd(db, file, content, imported_by)
        )

        db.commit()
        return payload

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# --------------------
# Imports list (CSV + XLSX)
# --------------------
@router.get("/imports")
def list_imports(limit: int = Query(20, ge=1, le=200), db: Session = Depends(get_db)):
    csv_q = (
        db.query(RawOrangePpdImport)
        .order_by(RawOrangePpdImport.imported_at.desc(), RawOrangePpdImport.import_id.desc())
        .limit(limit)
        .all()
    )
    csv_rows = [
        {
            "import_id": it.import_id,
            "filename": it.filename,
            "row_count": it.row_count,
            "imported_by": it.imported_by,
            "imported_at": it.imported_at.isoformat() if it.imported_at else None,
            "kind": "CSV",
            "sheet_name": None,
        }
        for it in csv_q
    ]

    xlsx_q = db.execute(
        text("""
            SELECT import_id, filename, sheet_name, row_count, imported_by, imported_at
            FROM canonique.orange_ppd_excel_imports
            ORDER BY imported_at DESC, import_id DESC
            LIMIT :limit
        """),
        {"limit": limit},
    ).mappings().all()

    xlsx_rows = [
        {
            "import_id": r["import_id"],
            "filename": r.get("filename"),
            "row_count": r.get("row_count"),
            "imported_by": r.get("imported_by"),
            "imported_at": r["imported_at"].isoformat() if r.get("imported_at") else None,
            "kind": "XLSX",
            "sheet_name": r.get("sheet_name"),
        }
        for r in xlsx_q
    ]

    merged = csv_rows + xlsx_rows
    merged.sort(key=lambda x: (x.get("imported_at") or ""), reverse=True)
    return merged[:limit]


@router.get("/excel-imports")
def list_excel_imports(limit: int = Query(20, ge=1, le=200), db: Session = Depends(get_db)):
    rows = db.execute(
        text("""
            SELECT import_id, filename, sheet_name, row_count, imported_by, imported_at
            FROM canonique.orange_ppd_excel_imports
            ORDER BY imported_at DESC, import_id DESC
            LIMIT :limit
        """),
        {"limit": limit},
    ).mappings().all()
    return [dict(r) for r in rows]


# --------------------
# PPD options (CSV only)
# --------------------
@router.get("/ppd-options")
def ppd_options(import_id: str | None = Query(None), db: Session = Depends(get_db)):
    import_id = _resolve_import_id(db, import_id)
    if not import_id:
        return []
    rows = db.execute(
        text("""
            SELECT DISTINCT NULLIF(BTRIM(numero_ppd),'') AS ppd
            FROM canonique.orange_ppd_rows
            WHERE import_id = :import_id
              AND NULLIF(BTRIM(numero_ppd),'') IS NOT NULL
            ORDER BY 1;
        """),
        {"import_id": import_id},
    ).all()
    return [r[0] for r in rows if r and r[0]]


# --------------------
# Helper: check si un import_id est XLSX
# --------------------
def _is_xlsx_import(db: Session, import_id: str | None) -> bool:
    if not import_id:
        return False
    return bool(
        db.execute(
            text("""
                SELECT 1
                FROM canonique.orange_ppd_excel_imports
                WHERE import_id = :import_id
                LIMIT 1
            """),
            {"import_id": import_id},
        ).scalar()
    )


# --------------------
# Compare (CSV ou XLSX)
# --------------------
@router.get("/compare")
def compare_orange_ppd(
    import_id: str | None = Query(default=None),
    ppd: str | None = Query(default=None),
    only_mismatch: bool = Query(default=False),
    db: Session = Depends(get_db),
):

    # ------------------------------------------------------------ XLSX
    if _is_xlsx_import(db, import_id):

        rows = db.execute(
            text("""
                SELECT
                  import_id,
                  n_cac        AS num_ot,
                  numero_ppd_orange,
                  releve,
                  nds,
                  facturation_orange_ht,
                  facturation_orange_ttc,
                  facturation_kyntus_ht,
                  facturation_kyntus_ttc,
                  diff_ht,
                  diff_ttc,
                  a_verifier,
                  true AS ot_existant,

                  CASE
                    WHEN match_found THEN 'OK'
                    ELSE 'ABSENT_PIDI'
                  END AS statut_croisement,

                  match_found AS croisement_complet,
                  reason

                FROM canonique.v_orange_ppd_excel_compare_releve

                WHERE import_id = :import_id
                  AND (:ppd IS NULL OR numero_ppd_orange = :ppd)
                  AND (:only_mismatch = FALSE OR a_verifier = TRUE)

                ORDER BY n_cac, releve
            """),
            {
                "import_id": import_id,
                "ppd": ppd,
                "only_mismatch": only_mismatch,
            },
        ).mappings().all()

        return [dict(r) for r in rows]

    # ------------------------------------------------------------ CSV

    rows = db.execute(
        text("""
            SELECT
              v.import_id,
              v.num_ot,
              v.numero_ppd_orange,

              COALESCE(v.facturation_orange_ht,  0)::numeric(12,2) AS facturation_orange_ht,
              COALESCE(v.facturation_orange_ttc, 0)::numeric(12,2) AS facturation_orange_ttc,

              COALESCE(v.facturation_kyntus_ht,  0)::numeric(12,2) AS facturation_kyntus_ht,
              COALESCE(v.facturation_kyntus_ttc, 0)::numeric(12,2) AS facturation_kyntus_ttc,

              COALESCE(v.diff_ht,  0)::numeric(12,2) AS diff_ht,
              COALESCE(v.diff_ttc, 0)::numeric(12,2) AS diff_ttc,

              v.a_verifier,
              v.ot_existant,

              CASE
                WHEN COALESCE(v.statut_croisement, '') IN ('', 'OT_INEXISTANT')
                  THEN 'ABSENT_PIDI'
                ELSE v.statut_croisement
              END AS statut_croisement,

              v.croisement_complet,

              CASE
                WHEN COALESCE(v.reason, '') IN ('', 'OT_INEXISTANT', 'CROISEMENT_INCOMPLET')
                 AND COALESCE(v.facturation_kyntus_ht, 0) = 0
                  THEN 'ABSENT_PIDI'
                ELSE COALESCE(v.reason, 'ABSENT_PIDI')
              END AS reason,

              ARRAY_REMOVE(
                ARRAY_AGG(DISTINCT NULLIF(BTRIM(r.nd), '')),
                NULL
              ) AS nds,

              MIN(NULLIF(BTRIM(r.nd), '')) AS releve

            FROM canonique.v_orange_ppd_compare v

            LEFT JOIN canonique.orange_ppd_rows r
              ON  r.import_id = v.import_id
              AND UPPER(regexp_replace(BTRIM(r.numero_ot), '\\s+', '', 'g'))
                = UPPER(regexp_replace(BTRIM(v.num_ot),   '\\s+', '', 'g'))

            WHERE (:import_id IS NULL OR v.import_id = :import_id)
              AND (:ppd IS NULL OR v.numero_ppd_orange = :ppd)
              AND (:only_mismatch = FALSE OR v.a_verifier = TRUE)

            GROUP BY
              v.import_id, v.num_ot, v.numero_ppd_orange,
              v.facturation_orange_ht,  v.facturation_orange_ttc,
              v.facturation_kyntus_ht,  v.facturation_kyntus_ttc,
              v.diff_ht,  v.diff_ttc,
              v.a_verifier, v.ot_existant, v.statut_croisement,
              v.croisement_complet, v.reason

            ORDER BY v.num_ot
        """),
        {"import_id": import_id, "ppd": ppd, "only_mismatch": only_mismatch},
    ).mappings().all()

    return [dict(r) for r in rows]


# --------------------
# Compare summary
# --------------------
@router.get("/compare-summary")
def compare_orange_ppd_summary(
    import_id: str | None = Query(default=None),
    ppd: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    _empty = {
        "orange_total_ht": 0, "orange_total_ttc": 0,
        "kyntus_total_ht": 0, "kyntus_total_ttc": 0,
        "ecart_ht": 0, "ecart_ttc": 0,
    }

    # ------------------------------------------------------------------ XLSX
    if _is_xlsx_import(db, import_id):
        sql = """
        WITH base AS (
          SELECT
            NULLIF(BTRIM(r.commande), '') AS commande_raw,
            COALESCE(r.montant_brut, 0)::numeric(12,2)   AS orange_ht,
            COALESCE(r.montant_majore, 0)::numeric(12,2) AS orange_ttc
          FROM canonique.orange_ppd_excel_rows r
          WHERE r.import_id = :import_id
            AND (:ppd IS NULL OR NULLIF(BTRIM(r.ppd_num), '') = :ppd)
        ),
        o AS (
          SELECT
            UPPER(regexp_replace(BTRIM(tok), '\\s+', '', 'g')) AS cac_key,
            SUM(b.orange_ht)::numeric(12,2)  AS orange_ht,
            SUM(b.orange_ttc)::numeric(12,2) AS orange_ttc
          FROM base b
          CROSS JOIN LATERAL
            regexp_split_to_table(COALESCE(b.commande_raw, ''), '[,;|\\n\\r\\t ]+') AS tok
          WHERE NULLIF(BTRIM(tok), '') IS NOT NULL
          GROUP BY 1
        ),
        p AS (
          SELECT
            UPPER(regexp_replace(BTRIM(p.n_cac), '\\s+', '', 'g')) AS cac_key,
            SUM(COALESCE(p.ht, 0))::numeric(12,2) AS pidi_ht,
            SUM(
              COALESCE(
                NULLIF(
                  REPLACE(regexp_replace(BTRIM(p.bordereau), '[^0-9,\\.\\-]', '', 'g'), ',', '.'),
                  ''
                )::numeric,
                0
              )
            )::numeric(12,2) AS pidi_bordereau
          FROM raw.pidi p
          WHERE NULLIF(BTRIM(p.n_cac), '') IS NOT NULL
          GROUP BY 1
        )
        SELECT
          COALESCE(SUM(o.orange_ht), 0)::numeric(12,2)                               AS orange_total_ht,
          COALESCE(SUM(o.orange_ttc), 0)::numeric(12,2)                              AS orange_total_ttc,
          COALESCE(SUM(COALESCE(p.pidi_ht, 0)), 0)::numeric(12,2)                   AS kyntus_total_ht,
          COALESCE(SUM(COALESCE(p.pidi_bordereau, 0)), 0)::numeric(12,2)            AS kyntus_total_ttc,
          (COALESCE(SUM(o.orange_ht), 0) - COALESCE(SUM(COALESCE(p.pidi_ht, 0)), 0))::numeric(12,2)         AS ecart_ht,
          (COALESCE(SUM(o.orange_ttc), 0) - COALESCE(SUM(COALESCE(p.pidi_bordereau, 0)), 0))::numeric(12,2) AS ecart_ttc
        FROM o
        LEFT JOIN p ON p.cac_key = o.cac_key
        """
        row = db.execute(text(sql), {"import_id": import_id, "ppd": ppd}).mappings().first()
        return dict(row) if row else _empty

    # ------------------------------------------------------------------ CSV
    row = db.execute(
        text("""
            SELECT
              COALESCE(SUM(facturation_orange_ht), 0)::numeric(12,2)  AS orange_total_ht,
              COALESCE(SUM(facturation_orange_ttc), 0)::numeric(12,2) AS orange_total_ttc,
              COALESCE(SUM(facturation_kyntus_ht), 0)::numeric(12,2)  AS kyntus_total_ht,
              COALESCE(SUM(facturation_kyntus_ttc), 0)::numeric(12,2) AS kyntus_total_ttc,
              (COALESCE(SUM(facturation_orange_ht), 0)
                - COALESCE(SUM(facturation_kyntus_ht), 0))::numeric(12,2)  AS ecart_ht,
              (COALESCE(SUM(facturation_orange_ttc), 0)
                - COALESCE(SUM(facturation_kyntus_ttc), 0))::numeric(12,2) AS ecart_ttc
            FROM canonique.v_orange_ppd_compare
            WHERE (:import_id IS NULL OR import_id = :import_id)
              AND (:ppd IS NULL OR numero_ppd_orange = :ppd)
        """),
        {"import_id": import_id, "ppd": ppd},
    ).mappings().first()
    return dict(row) if row else _empty

@router.get("/compare-tree")
def compare_orange_ppd_tree(
    import_id: str | None = Query(default=None),
    ppd: str | None = Query(default=None),
    only_mismatch: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    """
    Retourne une structure hiérarchique:
      CAC (commande) -> Relevé (comment_acqui_rejet) -> ND
    Montants Kyntus TTC = somme(bordereau parsé)  [Option A]
    Ne touche pas aux vues existantes, ne casse pas /compare.
    """

    # -------------------- helpers local
    def _clean_key(s: str | None) -> str | None:
        if not s:
            return None
        return re.sub(r"\s+", "", str(s).strip()).upper() or None

    def _norm_releve_key(s: str | None) -> str | None:
        # même logique que ta vue: trim + ltrim 0 + enlever non alnum + upper
        if not s:
            return None
        x = str(s).strip()
        x = x.lstrip("0")
        x = re.sub(r"[^0-9A-Za-z]", "", x)
        x = x.upper()
        return x or None

    # -------------------- XLSX ONLY (ton besoin actuel correspond à l’Excel)
    if not _is_xlsx_import(db, import_id):
        raise HTTPException(status_code=400, detail="compare-tree est prévu pour les imports XLSX (Excel).")

    # 1) Niveau CAC+Relevé depuis la vue existante (déjà comparée)
    base_rows = db.execute(
        text("""
            SELECT
              import_id,
              n_cac        AS num_ot,
              releve,
              numero_ppd_orange,
              facturation_orange_ht,
              facturation_orange_ttc,
              facturation_kyntus_ht,
              facturation_kyntus_ttc,
              diff_ht,
              diff_ttc,
              a_verifier,
              reason,
              nds
            FROM canonique.v_orange_ppd_excel_compare_releve
            WHERE import_id = :import_id
              AND (:ppd IS NULL OR numero_ppd_orange = :ppd)
              AND (:only_mismatch = FALSE OR a_verifier = TRUE)
            ORDER BY n_cac, releve
        """),
        {"import_id": import_id, "ppd": ppd, "only_mismatch": only_mismatch},
    ).mappings().all()

    if not base_rows:
        return []

    # 2) Détails ND depuis PIDI: group by CAC+Relevé+ND
    #    TTC Kyntus = somme(bordereau parsé) (Option A)
    nd_rows = db.execute(
        text(r"""
            SELECT
              upper(regexp_replace(btrim(p.n_cac), '\s+', '', 'g')) AS cac_key,
              upper(regexp_replace(ltrim(NULLIF(btrim(p.comment_acqui_rejet), ''), '0'), '[^0-9A-Za-z]', '', 'g')) AS releve_key,
              NULLIF(btrim(p.n_cac), '') AS n_cac,
              NULLIF(btrim(p.comment_acqui_rejet), '') AS releve,
              NULLIF(btrim(p.nd), '') AS nd,

              sum(coalesce(p.ht, 0))::numeric(12,2) AS pidi_ht,

              sum(
                coalesce(
                  nullif(
                    replace(regexp_replace(btrim(p.bordereau), '[^0-9,.\-]', '', 'g'), ',', '.'),
                    ''
                  )::numeric,
                  0
                )
              )::numeric(12,2) AS pidi_ttc

            FROM raw.pidi p
            WHERE NULLIF(btrim(p.n_cac), '') IS NOT NULL
              AND NULLIF(btrim(p.comment_acqui_rejet), '') IS NOT NULL
            GROUP BY 1,2,3,4,5
        """)
    ).mappings().all()

    # index: (cac_key, releve_key) -> list(nd items)
    nd_index: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for r in nd_rows:
        ck = _clean_key(r["n_cac"]) or (r["cac_key"] or "")
        rk = _norm_releve_key(r["releve"]) or (r["releve_key"] or "")
        if not ck or not rk:
            continue
        nd_index.setdefault((ck, rk), []).append(
            {
                "nd": r["nd"],
                "facturation_kyntus_ht": float(r["pidi_ht"] or 0),
                "facturation_kyntus_ttc": float(r["pidi_ttc"] or 0),
            }
        )

    # 3) Build tree: CAC -> Relevé -> ND
    #    On agrège aussi un total CAC à partir des enfants (Orange/Kyntus/diff)
    cac_tree: dict[str, dict[str, Any]] = {}

    for r in base_rows:
        cac = str(r["num_ot"] or "").strip()
        releve = str(r["releve"] or "").strip()
        if not cac:
            continue

        cac_key = _clean_key(cac) or ""
        releve_key = _norm_releve_key(releve) or ""

        # init CAC node
        node = cac_tree.get(cac)
        if not node:
            node = {
                "num_ot": cac,
                "numero_ppd_orange": r.get("numero_ppd_orange"),
                "facturation_orange_ht": 0.0,
                "facturation_orange_ttc": 0.0,
                "facturation_kyntus_ht": 0.0,
                "facturation_kyntus_ttc": 0.0,
                "diff_ht": 0.0,
                "diff_ttc": 0.0,
                "a_verifier": False,
                "children": [],
            }
            cac_tree[cac] = node

        orange_ht = float(r.get("facturation_orange_ht") or 0)
        orange_ttc = float(r.get("facturation_orange_ttc") or 0)
        kyntus_ht = float(r.get("facturation_kyntus_ht") or 0)
        kyntus_ttc = float(r.get("facturation_kyntus_ttc") or 0)
        diff_ht = float(r.get("diff_ht") or 0)
        diff_ttc = float(r.get("diff_ttc") or 0)
        a_verif = bool(r.get("a_verifier") or False)

        # children ND list for this CAC+Relevé
        nd_children = nd_index.get((cac_key, releve_key), [])

        releve_node = {
            "releve": releve or None,
            "numero_ppd_orange": r.get("numero_ppd_orange"),
            "facturation_orange_ht": orange_ht,
            "facturation_orange_ttc": orange_ttc,
            "facturation_kyntus_ht": kyntus_ht,
            "facturation_kyntus_ttc": kyntus_ttc,
            "diff_ht": diff_ht,
            "diff_ttc": diff_ttc,
            "a_verifier": a_verif,
            "reason": r.get("reason"),
            "nds": r.get("nds"),
            "children": nd_children,
        }

        node["children"].append(releve_node)

        # roll-up totals at CAC level
        node["facturation_orange_ht"] += orange_ht
        node["facturation_orange_ttc"] += orange_ttc
        node["facturation_kyntus_ht"] += kyntus_ht
        node["facturation_kyntus_ttc"] += kyntus_ttc
        node["diff_ht"] += diff_ht
        node["diff_ttc"] += diff_ttc
        node["a_verifier"] = node["a_verifier"] or a_verif

    # 4) return list sorted
    out = list(cac_tree.values())
    out.sort(key=lambda x: (x.get("num_ot") or ""))
    return out
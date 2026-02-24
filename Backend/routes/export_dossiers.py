# backend/routes/export_dossiers.py
from __future__ import annotations

from io import BytesIO
from typing import Optional
import re

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from database.connection import get_db

router = APIRouter(prefix="/api/dossiers", tags=["dossiers-export"])


def _autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row, 2000) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def _excel_safe(v):
    if v is None:
        return None
    try:
        tz = getattr(v, "tzinfo", None)
        if tz is not None and callable(getattr(v, "replace", None)):
            return v.replace(tzinfo=None)
    except Exception:
        pass
    return v


def _get_status_color(statut: str) -> str:
    colors = {
        "FACTURABLE": "C6EFCE",
        "NON_FACTURABLE": "FFC7CE",
        "A_VERIFIER": "FFEB9C",
        "CONDITIONNEL": "FFF2CC",
        "OK": "C6EFCE",
        "ABSENT_PRAXEDO": "FFEB9C",
        "ABSENT_PIDI": "FFC7CE",
        "INCONNU": "E7E6E6",
    }
    return colors.get(statut, "FFFFFF")


def _clean_filename(filename: str) -> str:
    """Nettoie le nom de fichier pour enlever les caractères problématiques"""
    # Remplacer les caractères non alphanumériques par des underscores
    filename = re.sub(r'[^\w\-_.]', '_', filename)
    # S'assurer que l'extension .xlsx est présente
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    return filename


@router.get("/export.xlsx")
def export_dossiers_xlsx(
    db: Session = Depends(get_db),
    statut_final: str | None = Query(None, alias="statut"),
    statut_croisement: str | None = Query(None, alias="croisement"),
    ppd: str | None = Query(None),
    limit: int = Query(50000, ge=1, le=200000),
):
    try:
        sql = text(
            """
            SELECT
                d.ot_key,
                d.nd_global,
                d.numero_ppd,
                d.attachement_valide,
                d.activite_code,
                d.produit_code,
                d.code_cible,
                d.code_cloture_code,
                d.mode_passage,
                d.type_site_terrain,
                d.type_pbo_terrain,

                d.regle_code,
                d.libelle_regle,
                d.statut_facturation,
                d.statut_final,
                d.motif_verification,
                d.is_previsite,
                d.statut_croisement,

                d.statut_praxedo,
                d.statut_pidi,
                d.date_planifiee,
                d.generated_at,
                d.technicien,

                d.liste_articles AS articles_pidi_brut,

                CASE
                    WHEN d.liste_articles IS NOT NULL
                    THEN array_to_string(canonique.parse_articles_piditext(d.liste_articles), ' | ')
                    ELSE NULL
                END AS articles_app,

                CASE
                    WHEN d.regle_articles_attendus IS NOT NULL THEN
                        array_to_string(
                            ARRAY(
                                SELECT jsonb_array_elements_text(
                                    CASE
                                        WHEN jsonb_typeof(d.regle_articles_attendus) = 'array'
                                        THEN d.regle_articles_attendus
                                        WHEN jsonb_typeof(d.regle_articles_attendus) = 'object'
                                             AND d.regle_articles_attendus ? 'articles'
                                        THEN d.regle_articles_attendus->'articles'
                                        ELSE '[]'::jsonb
                                    END
                                )
                            ),
                            ' | '
                        )
                    ELSE NULL
                END AS articles_attendus_regle,
                
                -- NOUVEAUX CHAMPS
                d.palier,
                d.palier_phrase,
                d.compte_rendu,
                CASE
                    WHEN d.evenements IS NOT NULL THEN left(d.evenements, 300)
                    ELSE NULL
                END AS evenements_extrait

            FROM canonique.v_dossier_facturable d
            WHERE 1=1
              AND (:statut IS NULL OR d.statut_final = :statut)
              AND (:croisement IS NULL OR d.statut_croisement = :croisement)
              AND (:ppd IS NULL OR COALESCE(d.numero_ppd,'') ILIKE '%' || :ppd || '%')
            ORDER BY d.generated_at DESC NULLS LAST
            LIMIT :limit
            """
        )

        rows = db.execute(
            sql,
            {"statut": statut_final, "croisement": statut_croisement, "ppd": ppd, "limit": limit},
        ).mappings().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Dossiers"

        headers = [
            "OT", "ND", "PPD", "Attachement", "Act", "Prod", "Code cible", "Clôture",
            "Terrain", "Type site", "Type PBO",
            "Règle", "Libellé règle", "Statut règle",
            "Statut final", "Motif vérif", "Prévisite",
            "Croisement", "Praxedo", "PIDI",
            "Planifiée", "Généré le", "Technicien",
            "Articles PIDI (brut)", "Articles APP (parsés)", "Articles attendus (règle)",
            "Palier", "Palier phrase", "Compte-rendu", "Evenements (extrait)"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.row_dimensions[1].height = 30

        for r in rows:
            row_data = [
                _excel_safe(r.get("ot_key")),
                _excel_safe(r.get("nd_global")),
                _excel_safe(r.get("numero_ppd")),
                _excel_safe(r.get("attachement_valide")),
                _excel_safe(r.get("activite_code")),
                _excel_safe(r.get("produit_code")),
                _excel_safe(r.get("code_cible")),
                _excel_safe(r.get("code_cloture_code")),
                _excel_safe(r.get("mode_passage")),
                _excel_safe(r.get("type_site_terrain")),
                _excel_safe(r.get("type_pbo_terrain")),
                _excel_safe(r.get("regle_code")),
                _excel_safe(r.get("libelle_regle")),
                _excel_safe(r.get("statut_facturation")),
                _excel_safe(r.get("statut_final")),
                _excel_safe(r.get("motif_verification")),
                "Oui" if r.get("is_previsite") else "Non",
                _excel_safe(r.get("statut_croisement")),
                _excel_safe(r.get("statut_praxedo")),
                "Validé" if r.get("statut_pidi") else "Non envoyé",
                _excel_safe(r.get("date_planifiee")),
                _excel_safe(r.get("generated_at")),
                _excel_safe(r.get("technicien")),
                _excel_safe(r.get("articles_pidi_brut")),
                _excel_safe(r.get("articles_app")),
                _excel_safe(r.get("articles_attendus_regle")),
                _excel_safe(r.get("palier")),
                _excel_safe(r.get("palier_phrase")),
                _excel_safe(r.get("compte_rendu")),
                _excel_safe(r.get("evenements_extrait")),
            ]

            row_num = ws.max_row + 1
            ws.append(row_data)

            if r.get("statut_final"):
                ws.cell(row=row_num, column=15).fill = PatternFill(
                    start_color=_get_status_color(r["statut_final"]),
                    end_color=_get_status_color(r["statut_final"]),
                    fill_type="solid",
                )

            if r.get("statut_croisement"):
                ws.cell(row=row_num, column=18).fill = PatternFill(
                    start_color=_get_status_color(r["statut_croisement"]),
                    end_color=_get_status_color(r["statut_croisement"]),
                    fill_type="solid",
                )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize_columns(ws)

        # Sauvegarder dans un buffer
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        # Construire le nom de fichier
        filename_parts = ["dossiers"]
        if statut_final:
            filename_parts.append(statut_final)
        if statut_croisement:
            filename_parts.append(statut_croisement)
        if ppd:
            # Nettoyer le PPD pour éviter les caractères problématiques
            clean_ppd = re.sub(r'[^\w\-]', '_', ppd)
            filename_parts.append(f"PPD_{clean_ppd}")
        
        filename = "_".join(filename_parts) + ".xlsx"
        filename = _clean_filename(filename)

        # CRITIQUE: Configurer correctement les headers
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Access-Control-Expose-Headers": "Content-Disposition",  # Important pour le frontend
        }

        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    
    except Exception as e:
        # Log l'erreur
        print(f"❌ Erreur lors de l'export Excel: {str(e)}")
        # Retourner une réponse d'erreur
        return JSONResponse(
            status_code=500,
            content={"detail": f"Erreur lors de l'export Excel: {str(e)}"}
        )
# Backend/routes/dossiers.py
from __future__ import annotations

import io
import json
import re
from typing import Any

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import case, or_
from sqlalchemy.orm import Session

from database.connection import get_db
from models.dossiers_facturable import VDossierFacturable
from schemas.dossier_facturable import DossierFacturable

router = APIRouter(prefix="/api/dossiers", tags=["dossiers"])


# ---------------------------
# Helpers export / formatting
# ---------------------------

_TOKEN_RE = re.compile(r"\b[A-Z]{2,}[A-Z0-9]{0,12}\b")


def _excel_cell(v: Any) -> str:
    """Convertit n'importe quelle valeur (jsonb, array, datetime, etc.) en string propre pour Excel."""
    if v is None:
        return ""
    if isinstance(v, (list, tuple, set)):
        return " | ".join(_excel_cell(x) for x in v)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    # datetime/date -> isoformat si possible
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat(sep=" ")
        except TypeError:
            return v.isoformat()
    return str(v)


def _extract_pidi_tokens(liste_articles: str | None) -> str:
    """Renvoie une string 'TOKEN1 | TOKEN2 | ...' comme dans l'UI chips."""
    if not liste_articles:
        return ""
    s = str(liste_articles).upper()
    matches = _TOKEN_RE.findall(s)
    toks = []
    for t in matches:
        t = t.strip()
        if not t:
            continue
        if t in ("PIDI", "BRUT"):
            continue
        toks.append(t)
    # unique en gardant l'ordre
    seen = set()
    out = []
    for t in toks:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return " | ".join(out)


def _base_query(
    db: Session,
    q: str | None,
    statut: str | None,
    croisement: str | None,
    ppd: str | None,
):
    qs = db.query(VDossierFacturable)

    if q:
        needle = q.strip()
        if needle:
            qs = qs.filter(
                or_(
                    VDossierFacturable.ot_key.ilike(f"%{needle}%"),
                    VDossierFacturable.nd_global.ilike(f"%{needle}%"),
                    VDossierFacturable.key_match.ilike(f"%{needle}%"),
                )
            )

    if statut:
        qs = qs.filter(VDossierFacturable.statut_final == statut)

    if croisement:
        qs = qs.filter(VDossierFacturable.statut_croisement == croisement)

    if ppd:
        needle_ppd = ppd.strip()
        if needle_ppd:
            qs = qs.filter(VDossierFacturable.numero_ppd.ilike(f"%{needle_ppd}%"))

    qs = qs.order_by(
        case(
            (VDossierFacturable.statut_final == "FACTURABLE", 1),
            (VDossierFacturable.statut_final == "CONDITIONNEL", 2),
            (VDossierFacturable.statut_final == "NON_FACTURABLE", 3),
            (VDossierFacturable.statut_final == "A_VERIFIER", 4),
            else_=5,
        ).asc(),
        case(
            (VDossierFacturable.motif_verification == "CROISEMENT_INCOMPLET", 1),
            else_=2,
        ).asc(),
        VDossierFacturable.generated_at.desc().nullslast(),
        VDossierFacturable.key_match.asc(),
    )

    return qs


# ---------------------------
# API list
# ---------------------------

@router.get("/", response_model=list[DossierFacturable])
def get_dossiers(
    q: str | None = None,
    statut: str | None = None,
    croisement: str | None = None,
    ppd: str | None = None,
    limit: int = Query(50, ge=1, le=5000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    qs = _base_query(db, q=q, statut=statut, croisement=croisement, ppd=ppd)
    return qs.limit(limit).offset(offset).all()


# ---------------------------
# API export XLSX (avec articles)
# ---------------------------

@router.get("/export.xlsx")
def export_dossiers_xlsx(
    q: str | None = None,
    statut: str | None = None,
    croisement: str | None = None,
    ppd: str | None = None,
    # IMPORTANT : par défaut on exporte tout (pas 50)
    db: Session = Depends(get_db),
):
    qs = _base_query(db, q=q, statut=statut, croisement=croisement, ppd=ppd)
    rows = qs.all()

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "dossiers"

    # ✅ Colonnes exportées (ajoute ici tout ce que tu veux)
    headers = [
        "key_match",
        "ot_key",
        "nd_global",
        "numero_ppd",
        "statut_final",
        "statut_croisement",
        "motif_verification",
        "activite_code",
        "produit_code",
        "code_cible",
        "code_cloture_code",
        "attachement_valide",
        "is_previsite",
        "technicien",
        "date_planifiee",
        "date_cloture",
        "generated_at",
        # ✅ Articles
        "article_facturation_propose",  # terrain proposé
        "liste_articles",               # PIDI brut
        "pidi_tokens",                  # tokens extraits (chips)
        # (optionnel) champs JSON utiles
        "services",
        "prix_degressifs",
        "articles_optionnels",
        "type_branchement",
        "justificatifs",
        "documents_attendus",
        "pieces_facturation",
        "outils_depose",
        # ✅ AJOUTS demandés (fin de liste)
        "commentaire_technicien",
        "source_facturation",
        "force_plp",
        "add_tsfh",
    ]

    ws.append(headers)

    for r in rows:
        ws.append([
            _excel_cell(getattr(r, "key_match", None)),
            _excel_cell(getattr(r, "ot_key", None)),
            _excel_cell(getattr(r, "nd_global", None)),
            _excel_cell(getattr(r, "numero_ppd", None)),
            _excel_cell(getattr(r, "statut_final", None)),
            _excel_cell(getattr(r, "statut_croisement", None)),
            _excel_cell(getattr(r, "motif_verification", None)),
            _excel_cell(getattr(r, "activite_code", None)),
            _excel_cell(getattr(r, "produit_code", None)),
            _excel_cell(getattr(r, "code_cible", None)),
            _excel_cell(getattr(r, "code_cloture_code", None)),
            _excel_cell(getattr(r, "attachement_valide", None)),
            _excel_cell(getattr(r, "is_previsite", None)),
            _excel_cell(getattr(r, "technicien", None)),
            _excel_cell(getattr(r, "date_planifiee", None)),
            _excel_cell(getattr(r, "date_cloture", None)),
            _excel_cell(getattr(r, "generated_at", None)),
            # ✅ Articles
            _excel_cell(getattr(r, "article_facturation_propose", None)),
            _excel_cell(getattr(r, "liste_articles", None)),
            _extract_pidi_tokens(getattr(r, "liste_articles", None)),
            # JSON / arrays
            _excel_cell(getattr(r, "services", None)),
            _excel_cell(getattr(r, "prix_degressifs", None)),
            _excel_cell(getattr(r, "articles_optionnels", None)),
            _excel_cell(getattr(r, "type_branchement", None)),
            _excel_cell(getattr(r, "justificatifs", None)),
            _excel_cell(getattr(r, "documents_attendus", None)),
            _excel_cell(getattr(r, "pieces_facturation", None)),
            _excel_cell(getattr(r, "outils_depose", None)),
            # ✅ AJOUTS demandés (fin de ligne)
            _excel_cell(getattr(r, "commentaire_technicien", None)),
            _excel_cell(getattr(r, "source_facturation", None)),
            _excel_cell(getattr(r, "force_plp", None)),
            _excel_cell(getattr(r, "add_tsfh", None)),
        ])

    # un minimum d'auto-size
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(45, max(12, len(headers[col_idx - 1]) + 2))

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "dossiers_export.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )